import configparser
import logging
import os
import re

import anndata as ad
import matplotlib.pyplot as plt
import mygene
import numpy as np
import pandas as pd
import scanpy as sc
import seaborn as sns
from openpyxl import load_workbook
from tqdm import tqdm

from ..auto_extract.agent import Claude3, Openai, get_cell_type_embedding_by_llm
from ..auto_extract.parse_params import Params 

def major_vote_top_clusters(df, 
                        leiden_key='leiden', 
                        cell_type_key='cell_type_raw', 
                        dataset_key='Dataset', 
                        top_n=5):
    """
    Analyze cell clusters and return top cell types for each cluster.
    
    Parameters:
    -----------
    df : pandas.DataFrame
        Input DataFrame containing cluster and cell type information
    leiden_key : str, default='leiden'
        Column name for cluster identification
    cell_type_key : str, default='cell_type_raw'
        Column name for cell type information
    dataset_key : str, default='Dataset'
        Column name for dataset information
    top_n : int, default=5
        Number of top cell types to return for each cluster
        
    Returns:
    --------
    pandas.DataFrame
        DataFrame containing analysis results
    """
    # Initialize empty list for storing results
    results = []
    
    # Iterate through each cluster
    for cluster in df[leiden_key].unique():
        # Get data for current cluster
        cluster_df = df[df[leiden_key] == cluster]
        
        # Calculate counts for each cell type
        cell_type_counts = cluster_df[cell_type_key].value_counts()
        
        # Get top N cell types by count
        top_cell_types = cell_type_counts.head(top_n)
        
        # Calculate unique Dataset support for each top cell type
        for cell_type in top_cell_types.index:
            cell_type_df = cluster_df[cluster_df[cell_type_key] == cell_type]
            dataset_support = len(cell_type_df[dataset_key].unique())
            
            # Add results to list
            results.append({
                'cluster': cluster,
                'cell_type': cell_type,
                'cell_count': cell_type_counts[cell_type],
                'dataset_support': dataset_support,
                'percentage': (cell_type_counts[cell_type] / len(cluster_df) * 100).round(2)
            })
    
    # Convert results to DataFrame
    result_df = pd.DataFrame(results)
    
    # Sort by cluster and cell count
    try:
        result_df['cluster'] = result_df['cluster'].astype(int)
    except ValueError:
        pass
    result_df = result_df.sort_values(['cluster', 'cell_count'], ascending=[True, False])
    
    return result_df

def majority_voting_cluster_annotation(cell_type_by_leiden_df: pd.DataFrame,
                                       config_path: str = 'config.ini'):
                                       
    config = configparser.ConfigParser()
    config.read(config_path)

    if 'openai' in config['API']['TYPE']:
        claude_agent = Openai(pdf_path=None, config_path=config_path)
    elif 'claude' in config['API']['TYPE']:
        claude_agent = Claude3(pdf_path=None, config_path=config_path)
    else:
        raise ValueError(f"Model {config['API']['MODEL']} not supported.")
    
    # split df into chunks of 50 rows
    content_list = []
    for i in range(0, len(cell_type_by_leiden_df), 50):
        end_idx = min(i+50, len(cell_type_by_leiden_df))
        content_list.append(cell_type_by_leiden_df.iloc[i:end_idx, :].to_string())
    
    params = Params(config_path)
    
    all_annotation_dict = {}
    all_reasoning = ""
    print('Majority voting annotation in progress...')
    for content in content_list:
        major_voting_prompt = params.get_tool_prompt('MAJOR_VOTE_ANNOTATION_PROMPT') + content
        model_response = claude_agent._tool_retrieve(messages=[{"role": "user", "content": major_voting_prompt}])
        annotation_dict = params.parse_annotation_response(model_response, simple_annotation=True)
        all_annotation_dict.update(annotation_dict)
        all_reasoning += model_response + '\n'
    
    return all_annotation_dict, all_reasoning

def plot_multiple_cluster_composition(df, clusters=None, figsize=(15, 3.5), min_percentage=1, ncols=3):
    """
    Create visualizations of cell type composition for multiple clusters in a grid layout.
    
    Parameters:
    -----------
    df : pandas.DataFrame
        DataFrame containing columns: 'cluster', 'cell_type', 'cell_count', 
        'dataset_support', 'percentage'
    clusters : list, optional
        List of cluster IDs to plot. If None, plot all clusters
    figsize : tuple, default=(15, 3.5)
        Figure size per row (width, height)
    min_percentage : float, default=1
        Minimum percentage to display in the plots
    ncols : int, default=3
        Number of clusters to display per row
    """
    import matplotlib.pyplot as plt
    import seaborn as sns
    import numpy as np
    
    # Determine which clusters to plot
    if clusters is None:
        clusters = sorted(df['cluster'].unique())
    else:
        clusters = [c for c in clusters if c in df['cluster'].unique()]
    
    # Calculate number of rows needed
    nrows = (len(clusters) + ncols - 1) // ncols
    
    # Adjust figure size based on number of rows
    adjusted_height = figsize[1] * nrows
    fig = plt.figure(figsize=(figsize[0], adjusted_height))
    
    # Create grid for subplots with increased column padding
    gs = fig.add_gridspec(nrows, ncols * 3, 
                         width_ratios=np.tile([2, 1, 1], ncols),  # Increased last ratio for padding
                         hspace=0.5,  # Vertical spacing between rows
                         wspace=0.8)  # Increased horizontal spacing between columns
    
    for idx, cluster_id in enumerate(clusters):
        # Calculate row and column position
        row = idx // ncols
        col = (idx % ncols) * 3  # *3 because we have 3 columns per cluster
        
        # Get data for current cluster
        plot_df = df[df['cluster'] == cluster_id].copy()
        plot_df = plot_df[plot_df['percentage'] >= min_percentage]
        
        # Create subplots for this cluster with different widths
        ax1 = fig.add_subplot(gs[row, col])    # Takes 2 width units
        ax2 = fig.add_subplot(gs[row, col+1])  # Takes 1 width unit
        
        # Color palette
        colors = sns.color_palette("husl", n_colors=len(plot_df))
        
        # First subplot: Cell counts and percentages
        bars = ax1.barh(range(len(plot_df)), plot_df['percentage'], color=colors)
        
        # Add cell count annotations inside the bars
        for i, bar in enumerate(bars):
            width = bar.get_width()
            count = plot_df['cell_count'].iloc[i]
            count_text = f'{count:,}'
            text_position = max(width * 0.01, 0.5)
            
            ax1.text(text_position, i,
                    f'{count_text} ({plot_df["percentage"].iloc[i]:.1f}%)',
                    va='center',
                    color='black')
        
        # Second subplot: Dataset support
        bars2 = ax2.barh(range(len(plot_df)), plot_df['dataset_support'],
                        color=colors, alpha=0.7)
        
        # Add dataset support annotations
        for i, bar in enumerate(bars2):
            width = bar.get_width()
            dataset_count = plot_df['dataset_support'].iloc[i]
            ax2.text(width + 0.1, i,
                    str(dataset_count),
                    va='center')
        
        # Customize both subplots
        for ax in [ax1, ax2]:
            ax.set_yticks(range(len(plot_df)))
            ax.grid(True, axis='x', alpha=0.3)
        
        # Set y-labels
        ax1.set_yticklabels(plot_df['cell_type'])
        ax2.set_yticklabels([])
        
        # Set titles and labels
        ax1.set_title(f"Cluster {cluster_id}")
        ax1.set_xlabel('Percentage of Cells')
        ax2.set_xlabel('N Datasets')
        
        # Adjust x-axis limits for dataset support plot
        ax2.set_xlim(0, max(df['dataset_support']) * 1.15)  # Add 15% padding
    
    # Adjust layout with specific padding
    fig.tight_layout()  # Increase padding around the entire figure
    
    return fig

def convert_ensembl_to_symbol(adata: ad.AnnData) -> ad.AnnData:
    """
    Convert Ensembl IDs to gene symbols.
    """
    
    mg = mygene.MyGeneInfo()
    shape_before = adata.shape
    
    logging.info('Finding gene symbols for Ensembl IDs...')
    adata.var.index = adata.var.index.str.split('.').str[0]
    
    retry = 0
    while retry < 3:
        try:
            query_dict_list = mg.querymany(adata.var.index, scopes='ensembl.gene', fields='symbol')
            break
        except:
            retry += 1
            logging.warning('Failed to connect to MyGeneInfo. Retrying...')
    
    logging.info('Converting Ensembl IDs to gene symbols...')
    convert_dict = {}
    for x in query_dict_list:
        if x['query'] in convert_dict.keys():
            print('duplicated Ensembl IDs: '+x['query'])
        if 'symbol' in x:
            convert_dict[x['query']] = x['symbol']
        else:
            convert_dict[x['query']] = x['query']
    
    for ind, rows in tqdm(adata.var.iterrows()):
        if ind in convert_dict.keys():
            adata.var.loc[ind, 'gene_symbol'] = convert_dict[ind]
        else:
            adata.var.loc[ind, 'gene_symbol'] = ind
    
    adata_extracted = adata[:, adata.var['gene_symbol'] != adata.var.index].copy()
    adata_extracted.var = adata_extracted.var.reset_index(names='Ensembl_ID').set_index('gene_symbol', drop=True).copy()
    
    shape_after = adata_extracted.shape
    
    logging.info(f"Successfully Converted {shape_after[1]} Ensembl IDs to gene symbols.")
    logging.info(f"{shape_before[1] - shape_after[1]} Ensembl IDs were not converted.")
    
    return adata_extracted 

def summarise_marker_gene_from_sheet(marker_genes_sheet_content: str, config_path: str = 'config.ini') -> str:
    config = configparser.ConfigParser()
    config.read(config_path)

    if 'openai' in config['API']['TYPE']:
        claude_agent = Openai(pdf_path=None, config_path=config_path)
    elif 'claude' in config['API']['TYPE']:
        claude_agent = Claude3(pdf_path=None, config_path=config_path)
    else:
        raise ValueError(f"Model {config['API']['MODEL']} not supported.")
    
    params = Params(config_path)
    
    summary_marker_gene_prompt = params.get_tool_prompt('SUMMARY_MARKER_GENES_FROM_SHEET_PROMPT').replace('{marker_genes_sheet}',
                                            marker_genes_sheet_content)
                    
    marker_gene_response = claude_agent._tool_retrieve(messages=[{"role": "user", "content": summary_marker_gene_prompt}])
    return marker_gene_response

def extract_top_markers_from_excel(input_path, top_k=10, min_score=0, max_pval=0.05, config_path='config.ini'):
    """
    Extract top marker genes for each cell type (Sheet) from an Excel file
    
    Parameters:
    input_path: Path to the Excel file
    top_k: Number of genes to extract for each cell type
    min_score: Minimum score threshold (expression change)
    max_pval: Maximum p-value threshold
    
    Returns:
    Dictionary with cell types (sheet names) as keys and lists of top marker genes as values
    """
    # Check if file exists
    if not os.path.exists(input_path):
        print(f"Error: Input file {input_path} does not exist")
        return {}
        
    # Check if file is Excel format
    if not input_path.endswith(('.xlsx', '.xls')):
        print(f"Error: Input file {input_path} is not an Excel file")
        return {}
    
    # Get all sheets from Excel
    try:
        wb = load_workbook(input_path, read_only=True)
        sheet_names = wb.sheetnames
    except Exception as e:
        print(f"Error reading Excel file: {e}")
        return {}
    
    if not sheet_names:
        print("No sheets found in Excel file")
        return {}
    
    print(f"Found {len(sheet_names)} sheets, extracting top markers...")
    
    # Create result dictionary
    cell_type_markers = {}
    
    if len(sheet_names) == 1:
        print(f"Only one sheet found: {sheet_names[0]}")
        sheet_content = pd.read_excel(input_path, sheet_name=sheet_names[0]).to_string()
        marker_gene_response = summarise_marker_gene_from_sheet(sheet_content, config_path)
        cell_type_markers['direct_parse'] = marker_gene_response
        return cell_type_markers
    
    # Process each sheet
    for sheet_name in sheet_names:
        try:
            # Read sheet data
            df = pd.read_excel(input_path, sheet_name=sheet_name)
            
            if df.empty:
                print(f"  Skipping empty sheet: {sheet_name}")
                continue
            
            # Auto-identify columns
            gene_col = identify_gene_column(df)
            # score_col = identify_score_column(df)
            # pval_col = identify_pvalue_column(df)
            
            if not gene_col:
                print(f"  Unable to identify gene column in sheet {sheet_name}, skipping")
                continue
            
            # Create a working copy for filtering and sorting
            filtered_df = df.copy()
            
            # # Apply filtering conditions
            # if score_col and min_score > 0:
            #     if pd.api.types.is_numeric_dtype(filtered_df[score_col]):
            #         filtered_df = filtered_df[filtered_df[score_col].abs() >= min_score]
            
            # if pval_col and max_pval < 1:
            #     if pd.api.types.is_numeric_dtype(filtered_df[pval_col]):
            #         filtered_df = filtered_df[filtered_df[pval_col] <= max_pval]
            
            # # Sort by score
            # if score_col:
            #     # Sort by absolute value
            #     filtered_df['abs_score'] = filtered_df[score_col].abs()
            #     filtered_df = filtered_df.sort_values(by='abs_score', ascending=False)
            #     if 'abs_score' in filtered_df.columns:
            #         filtered_df = filtered_df.drop(columns=['abs_score'])
            
            # Extract top K genes
            top_df = filtered_df.head(top_k)
            
            if top_df.empty:
                print(f"  No genes meeting criteria in sheet {sheet_name}")
                continue
            
            # Get gene list
            top_genes = top_df[gene_col].tolist()
            
            # Clean gene names (remove NaN etc.)
            top_genes = [str(gene).strip() for gene in top_genes if str(gene).strip() and str(gene).lower() != 'nan']
            
            # Add to result dictionary
            cell_type_markers[sheet_name] = top_genes
            
            print(f"  {sheet_name}: Extracted {len(top_genes)} markers")
            
        except Exception as e:
            print(f"  Error processing sheet {sheet_name}: {e}")
    
    return cell_type_markers

def format_markers_dict(markers_dict):
    """Format markers dictionary as 'celltype: marker1, marker2, ...' list"""
    formatted_list = []
    
    for cell_type, markers in markers_dict.items():
        markers_str = ", ".join(markers)
        formatted_list.append(f"{cell_type}: {markers_str}")
    
    return formatted_list

def identify_gene_column(df):
    """Identify column containing gene IDs/names"""
    # Common gene column names
    gene_candidates = ['gene', 'gene_name', 'gene_id', 'symbol', 'genesymbol', 'feature', 'genes']
    
    # First check column names
    for col in df.columns:
        if col.lower() in gene_candidates or any(gc in col.lower() for gc in gene_candidates):
            return col
    
    # If no explicit column name, check if first column contains gene ID patterns
    if len(df.columns) > 0:
        first_col = df.columns[0]
        sample_values = df[first_col].head(10).astype(str)
        
        # Check for common gene ID formats
        gene_patterns = [
            r'^ENS[A-Z]*G\d+', # Ensembl gene ID
            r'^[A-Z0-9]+_[A-Z0-9]+', # Other common formats
            r'^[A-Z][A-Z0-9]+$' # Gene symbol format (like CD3D, GAPDH etc.)
        ]
        
        for pattern in gene_patterns:
            if sample_values.str.match(pattern).any():
                return first_col
    
    # Default to first column
    if len(df.columns) > 0:
        return df.columns[0]
    
    return None

def identify_score_column(df):
    """Identify column containing expression change scores"""
    # Common score column names
    score_candidates = ['avg_log', 'avg_logfc', 'logfc', 'log2fc', 'log2_fc', 
                        'log2foldchange', 'log2_fold_change', 'fc', 'fold_change',
                        'score', 'effect', 'expr', 'expression', 'lfc']
    
    # Check column names
    for col in df.columns:
        col_lower = col.lower()
        if col_lower in score_candidates or any(sc in col_lower for sc in score_candidates):
            return col
    
    # If no explicit column name, look for numeric columns
    numeric_cols = [col for col in df.columns if pd.api.types.is_numeric_dtype(df[col])]
    
    # Exclude potential p-value columns
    pval_patterns = ['p', 'pval', 'p_val', 'p.val', 'padj', 'p_adj', 'fdr']
    numeric_cols = [col for col in numeric_cols if not any(p in col.lower() for p in pval_patterns)]
    
    if numeric_cols:
        # Return first non-p-value numeric column
        return numeric_cols[0]
    
    return None

def identify_pvalue_column(df):
    """Identify column containing p-values"""
    # Common p-value column names
    pval_candidates = ['padj', 'p_adj', 'adj_pval', 'p.val.adj', 'fdr', 'q_value', 'qval', 
                       'pval', 'p_val', 'p.val', 'p-value', 'p_value', 'p.value']
    
    # Prioritize adjusted p-value columns
    adj_pval_candidates = ['padj', 'p_adj', 'adj_pval', 'p.val.adj', 'fdr', 'q_value', 'qval']
    
    # First try to find adjusted p-value
    for col in df.columns:
        col_lower = col.lower()
        if col_lower in adj_pval_candidates or any(pc in col_lower for pc in adj_pval_candidates):
            return col
    
    # Then try to find unadjusted p-value
    for col in df.columns:
        col_lower = col.lower()
        if col_lower in pval_candidates or any(pc in col_lower for pc in pval_candidates):
            return col
    
    # If no explicit column name, look for potential p-value columns (usually very small values)
    numeric_cols = [col for col in df.columns if pd.api.types.is_numeric_dtype(df[col])]
    for col in numeric_cols:
        non_zero_values = df[col][df[col] != 0]
        if not non_zero_values.empty:
            if non_zero_values.max() <= 1 and non_zero_values.min() >= 0:
                # Check for very small values, which typically indicate p-values
                if non_zero_values.min() < 0.1:
                    return col
    
    return None

def get_top_markers(input_path, top_k=10, min_score=0, max_pval=0.05, config_path='config.ini'):
    """
    Extract and return top marker genes for each cell type, formatted as 'celltype: marker list'
    
    Parameters:
    input_path: Path to Excel file
    top_k: Number of genes to extract for each cell type, default is 10
    min_score: Minimum score threshold, default is 0
    max_pval: Maximum p-value threshold, default is 0.05
    
    Returns:
    List of strings, each formatted as 'celltype: marker1, marker2, ...'
    """
    # Extract markers dictionary
    markers_dict = extract_top_markers_from_excel(input_path, top_k, min_score, max_pval, config_path)
    
    if 'direct_parse' in markers_dict.keys():
        return [markers_dict['direct_parse']]
    
    # Format as required list format
    return format_markers_dict(markers_dict)

if __name__ == '__main__':
    adata = sc.read_h5ad('/home/wu/datb1/AutoExtractSingleCell/sample2/raw_data/raw_data.h5ad')
    adata_test = adata[:, adata.var.iloc[:10, :].index].copy()
    adata_test

    adata_extracted = convert_ensembl_to_symbol(adata_test)